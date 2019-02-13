import openpyxl,os

#创建一个新的excel

filedir = './file/'
wb = openpyxl.load_workbook('xlsxtotxt.xlsx')
sheet = wb.get_active_sheet()


for j in range(1,sheet.max_column+1):
	filedirname = filedir + str(j) + '.txt'
	for i in range(1,sheet.max_row+1):
		text = sheet.cell(row=i,column=j).value
		if text != None:
			with open(filedirname, 'a') as obj_f: 
				obj_f.write(text + '\n')



	

		






