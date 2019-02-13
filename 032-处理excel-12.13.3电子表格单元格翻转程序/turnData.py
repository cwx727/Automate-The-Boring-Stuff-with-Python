import openpyxl

#创建一个新的excel
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_active_sheet()

sheetData=[]

wb_turn = openpyxl.Workbook()
sheet_turn = wb_turn.get_active_sheet()

for i in range(1, sheet.max_row+1):
	sheetData.append([])
	for j in range(1, sheet.max_column+1):
		sheetData[i-1].append(sheet.cell(row=i, column=j).value)

for j in range(len(sheetData[0])):
	for i in range(len(sheetData)):
		sheet_turn.cell(row=j+1, column=i+1).value = sheetData[i][j]

wb_turn.save('exampleTurn.xlsx')
