import openpyxl

#创建一个新excel
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()

N = int(input('please input N:'))

def multiplicate(N):
	#首行首列值
	for i in range(1,N+1):
		sheet.cell(row=1, column=i+1).value = i
		sheet.cell(row=i+1, column=1).value = i
	#乘法表值
	for i in range(2,N+2):
		for j in range(2,N+2):
			sheet.cell(row =i, column =j).value = (i-1)*(j-1)

multiplicate(N)

#保存新excel内容
sheet.freeze_panes = 'A2' #冻结首行
wb.save('multiplicationTable.xlsx')
