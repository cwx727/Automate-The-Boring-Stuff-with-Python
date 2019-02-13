import openpyxl

M = int(input('Please input before which row you want to insert:'))
N = int(input('How many rows you want to insert:'))
filename = input('The filename is:')
#filename = 'row.xlsx'

#创建一个新的excel
wb_insert = openpyxl.Workbook()
sheet_insert = wb_insert.get_active_sheet()

def blankRowInsert(M,N,filename):
	#取原有文件中的内容
	wb = openpyxl.load_workbook(filename)
	sheet = wb.get_active_sheet()
	for j in range(1, sheet.max_column+1):
		#新excel的1到M行复制原有文件的内容
		for m in range(1,M):
			sheet_insert.cell(row=m,column=j).value = sheet.cell(row=m,column=j).value
		#新excel的M+N行往后复制原有文件的剩下的M+1行往后的内容
		for m in range(M+1,sheet.max_row+1):
			sheet_insert.cell(row=m+N-1,column=j).value = sheet.cell(row=m,column=j).value
		
blankRowInsert(M,N,filename)
wb_insert.save('blankRowInserter.xlsx')
